"""
Transaction CRUD router.
All routes are company-scoped — company_id comes from the JWT, never from the URL.
After writes, an SSE event is broadcast to all dashboard clients.
"""
import csv
import io
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Query, Request, status
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse

from app.core.deps import CompanyID, CurrentUser, DB
from app.schemas.transaction import (
    TransactionCreate,
    TransactionList,
    TransactionOut,
    TransactionUpdate,
)
from app.services import transactions as tx_service
from app.services.budgets import check_budget_alert

router = APIRouter(prefix="/transactions", tags=["transactions"])


def _get_broadcaster(request: Request):
    return request.app.state.broadcaster


def _get_bot(request: Request):
    return request.app.state.bot


@router.post("", response_model=TransactionOut, status_code=status.HTTP_201_CREATED)
async def create_transaction(
    request: Request,
    data: TransactionCreate,
    company_id: CompanyID,
    current_user: CurrentUser,
    db: DB,
) -> TransactionOut:
    out = await tx_service.create_transaction(company_id, current_user.id, data, db)

    # Check budget limits and fire alert if threshold crossed
    await check_budget_alert(
        company_id, data.category_id, db, bot=_get_bot(request)
    )

    # Broadcast SSE event to all connected dashboard clients
    await _get_broadcaster(request).broadcast(
        company_id,
        event_type="transaction.created",
        data=jsonable_encoder(out),
    )
    return out


@router.get("", response_model=TransactionList)
async def list_transactions(
    company_id: CompanyID,
    db: DB,
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    type_filter: str | None = Query(default=None, alias="type"),
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
) -> TransactionList:
    return await tx_service.list_transactions(
        company_id, db,
        page=page, limit=limit,
        type_filter=type_filter,
        date_from=date_from,
        date_to=date_to,
    )


async def _fetch_export_data(
    company_id, db, date_from, date_to, type_filter
) -> tuple[list, dict]:
    """Shared helper: fetch transactions + category name lookup."""
    from sqlalchemy import select
    from app.models.category import Category

    listing = await tx_service.list_transactions(
        company_id, db,
        page=1, limit=10_000,
        type_filter=type_filter,
        date_from=date_from,
        date_to=date_to,
    )

    result = await db.execute(
        select(Category.id, Category.name).where(Category.company_id == company_id)
    )
    cat_names: dict = {str(row.id): row.name for row in result.all()}

    return listing.items, cat_names


@router.get("/export/csv")
async def export_csv(
    company_id: CompanyID,
    db: DB,
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    type_filter: str | None = Query(default=None, alias="type"),
) -> StreamingResponse:
    """Export transactions as a UTF-8 CSV file."""
    items, cat_names = await _fetch_export_data(company_id, db, date_from, date_to, type_filter)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Sana", "Tur", "Summa (so'm)", "Kategoriya", "Izoh", "Manba", "Yaratilgan",
    ])
    for tx in items:
        writer.writerow([
            tx.date.strftime("%d.%m.%Y"),
            "Daromad" if tx.type == "income" else "Xarajat",
            str(tx.amount),
            cat_names.get(str(tx.category_id), "") if tx.category_id else "",
            tx.description or "",
            tx.source,
            tx.created_at.strftime("%d.%m.%Y %H:%M"),
        ])

    output.seek(0)
    filename = f"tranzaksiyalar_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]),   # BOM for correct Excel UTF-8 display
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/excel")
async def export_excel(
    company_id: CompanyID,
    db: DB,
    date_from: datetime | None = Query(default=None),
    date_to: datetime | None = Query(default=None),
    type_filter: str | None = Query(default=None, alias="type"),
) -> StreamingResponse:
    """Export transactions as an Excel (.xlsx) file with formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=501,
            detail="Excel eksport uchun openpyxl kutubxonasi kerak: pip install openpyxl",
        )

    items, cat_names = await _fetch_export_data(company_id, db, date_from, date_to, type_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tranzaksiyalar"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo-600
    center = Alignment(horizontal="center", vertical="center")
    income_fill = PatternFill("solid", fgColor="D1FAE5")   # emerald-100
    expense_fill = PatternFill("solid", fgColor="FEE2E2")  # rose-100
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Sana", "Tur", "Summa (so'm)", "Kategoriya", "Izoh", "Manba", "Yaratilgan"]
    col_widths = [14, 10, 18, 18, 30, 10, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, tx in enumerate(items, start=2):
        is_income = tx.type == "income"
        row_fill = income_fill if is_income else expense_fill

        values = [
            tx.date.strftime("%d.%m.%Y"),
            "Daromad" if is_income else "Xarajat",
            float(tx.amount),
            cat_names.get(str(tx.category_id), "") if tx.category_id else "",
            tx.description or "",
            tx.source,
            tx.created_at.strftime("%d.%m.%Y %H:%M"),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = border
            if col_idx == 3:  # amount column — number format
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # ── Summary row ───────────────────────────────────────────────────────────
    if items:
        sum_row = len(items) + 2
        ws.cell(row=sum_row, column=1, value="JAMI").font = Font(bold=True)
        sum_cell = ws.cell(
            row=sum_row, column=3,
            value=f"=SUMIF(B2:B{len(items)+1},\"Daromad\",C2:C{len(items)+1})-SUMIF(B2:B{len(items)+1},\"Xarajat\",C2:C{len(items)+1})",
        )
        sum_cell.font = Font(bold=True)
        sum_cell.number_format = '#,##0.00'

    # ── Freeze header, enable autofilter ─────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(items)+1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"tranzaksiyalar_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{tx_id}", response_model=TransactionOut)
async def get_transaction(
    tx_id: uuid.UUID, company_id: CompanyID, db: DB
) -> TransactionOut:
    tx = await tx_service.get_transaction(company_id, tx_id, db)
    return TransactionOut.model_validate(tx)


@router.patch("/{tx_id}", response_model=TransactionOut)
async def update_transaction(
    request: Request,
    tx_id: uuid.UUID,
    data: TransactionUpdate,
    company_id: CompanyID,
    db: DB,
) -> TransactionOut:
    out = await tx_service.update_transaction(company_id, tx_id, data, db)

    await _get_broadcaster(request).broadcast(
        company_id,
        event_type="transaction.updated",
        data=jsonable_encoder(out),
    )
    return out


@router.delete("/{tx_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_transaction(
    request: Request,
    tx_id: uuid.UUID,
    company_id: CompanyID,
    db: DB,
) -> None:
    await tx_service.soft_delete_transaction(company_id, tx_id, db)

    await _get_broadcaster(request).broadcast(
        company_id,
        event_type="transaction.deleted",
        data={"id": str(tx_id)},
    )
